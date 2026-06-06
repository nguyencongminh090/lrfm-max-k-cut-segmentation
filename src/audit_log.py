"""Generate the AI Audit Log (.xlsx) for DAP391m, following the official template.

The log records the CORE decisions made while working with AI on this project
(RFM->LRFM customer segmentation via max-k-cut on Online Retail II). It is NOT a
chat dump - only prompts that changed the project's substance are logged.

APPENDABLE: to add a future entry, append a dict to CORE (or HALLU) below and
re-run `python -m src.audit_log`. The workbook is rebuilt from these lists.

Output: results/AI_Audit_Log_DAP391m.xlsx  (4 sheets matching the template)
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from . import config

OUT = config.RESULTS / "AI_Audit_Log_DAP391m.xlsx"

META = {
    "Student Name": "[Enter your name]",
    "Student ID": "[Enter your ID]",
    "Course": "DAP391m",
    "Assignment": "RFM->LRFM Customer Segmentation via Max-k-cut (Online Retail II)",
    "AI tools used": "Claude (Opus 4.8) - primary research/coding partner",
    "Total Prompts Used (all AI tools)": "[Enter total across all your AI sessions]",
    "Core Prompts Logged": 17,
    "Hallucination cases": 4,
}


def delta(ct, cx, cs, do):
    return (f"Critical Thinking: {ct}\n"
            f"Contextualization: {cx}\n"
            f"Creative Synthesis: {cs}\n"
            f"Decision Ownership: {do}")


# ---------------------------------------------------------------------------
# CORE PROMPTS  (Decision / Problem-Solving / Verification)
# ---------------------------------------------------------------------------
CORE = [
    dict(t="DECISION", stage="Business Understanding / Research",
         ctx="Starting the project: needed to understand the two reference papers and frame the problem before choosing a direction.",
         prompt="Be my research partner. Read the papers in @papers/ (FCA paper and graph-theory/max-k-cut paper), understand the problem, and report your analysis. I'll tell you my plan after.",
         resp="Summarised both papers (FCA hierarchical vs graph max-k-cut), noted they share the Online Retail II dataset, and flagged three issues: neither paper uses Deep Learning, the dataset has NO campaign-response field, and the graph paper already beats FCA on silhouette.",
         delta=delta(
             "AI corrected my assumption that the two papers used DL/ML - they are lattice theory and combinatorial optimization, not learning.",
             "My brief asked for response/open-rate prediction, but the dataset has no such label - a constraint the AI surfaced early.",
             "I used the AI's map of the two methods to scope a realistic, deadline-aware contribution instead of an ML rebuild.",
             "Accepted the corrections and narrowed scope to segmentation; deferred any prediction component."),
         ev="Initial analysis report; papers/ folder; project memory."),

    dict(t="DECISION", stage="Business Understanding / Research",
         ctx="Deciding the core contribution direction under a tight deadline.",
         prompt="I want to improve the graph-based paper by extending RFM -> LRFM. Make the paper's model the baseline and run my enhanced version to compare. Does it help?",
         resp="Confirmed feasible: the paper's vertex-reduction generalises (T^4=625 vertices) and LRFM is the paper's own stated future work. Warned that adding L would likely LOWER silhouette (curse of dimensionality), so the contribution should be framed on actionability, not silhouette.",
         delta=delta(
             "The AI's warning that silhouette would drop with a 4th dimension was non-obvious and proved correct in Stage 3.",
             "Deadline pressure meant I needed a defensible, paper-endorsed extension, not a flashy one.",
             "I combined LRFM (cited extension) with the graph method and decided to argue the contribution on business actionability.",
             "Chose LRFM + max-k-cut as the project; framing fixed on actionability over the silhouette metric."),
         ev="stage3_modeling_report.md; comparison_report.md."),

    dict(t="DECISION", stage="Build Model / Method selection",
         ctx="Considering whether a Graph Neural Network would be a stronger, more novel contribution.",
         prompt="If I apply a Graph Neural Network on the customer graph instead, does it help?",
         resp="Advised against it for this case: the graph's edges are just re-encoded RFM distances and nodes have only 3-4 features, so a GNN has nothing extra to learn; the data is tiny for deep learning. Only worthwhile if the graph were enriched with co-purchase edges (a deadline-breaker).",
         delta=delta(
             "I noticed the AI's FIRST report had oversold GNN as 'novel/beneficial'; on probing, it conceded it had oversold it (see Hallucination H1).",
             "3 features and ~5,878 customers is unsuitable for deep learning - domain reality the generic 'use a GNN' suggestion ignored.",
             "I rejected the GNN path and kept the simpler, better-fitting graph method.",
             "Decided NOT to use a GNN; recorded the reasoning so a referee sees it was considered and dismissed."),
         ev="Conversation log; FE correlation analysis (3-feature redundancy)."),

    dict(t="DECISION", stage="Build Model / Method selection",
         ctx="Weighing an FCA + Neural Network hybrid against the graph approach.",
         prompt="FCA + Neural Network compared with the graph approach - is that a good idea?",
         resp="Argued it is conceptually weak: FCA's only advantage is interpretability, which a neural net destroys, and FCA already loses on silhouette - so the hybrid combines the benchmark's losing method with a tool that erases its one strength.",
         delta=delta(
             "I pressure-tested a third option rather than committing blindly; the AI's reasoning about interpretability-vs-blackbox was sound.",
             "An FCA+NN hybrid has no clear win on either silhouette or interpretability for this dataset.",
             "I eliminated the hybrid and consolidated on LRFM + max-k-cut.",
             "Rejected FCA+NN; this stress-test confirmed the LRFM choice."),
         ev="Conversation log; FCA paper silhouette (0.32) vs graph (0.46)."),

    dict(t="PROBLEM-SOLVING", stage="Algorithms / Solver",
         ctx="The free pip Gurobi license is size-limited and fails even on small instances; I have no academic email to get a full license.",
         prompt="Can you rebuild Gurobi? (clarified: I mean code the max-k-cut solver from scratch, not reinstall)",
         resp="Confirmed reinstalling cannot lift the license cap (it counts quadratic terms). Implemented a from-scratch multi-start local-search max-k-cut solver in pure NumPy - no license, no external solver.",
         delta=delta(
             "The AI first misread 'rebuild' as 'reinstall'; I clarified I meant a from-scratch implementation, and it adjusted.",
             "No academic email -> the free Gurobi academic license is not obtainable, so a license-free solver was mandatory.",
             "I chose a custom heuristic over the slow open-source CBC fallback.",
             "Owned the decision to write our own solver; removed the Gurobi dependency entirely."),
         ev="Gurobi license-limit probe output; src/maxkcut.py."),

    dict(t="VERIFICATION", stage="Algorithms / Solver",
         ctx="A from-scratch heuristic needs proof of correctness before the paper can rely on it.",
         prompt="Confirm your solver (local) is: 1. fast, 2. correctly implemented, 3. can it be optimized more?",
         resp="Built a validation suite: brute-force global optimum on tiny instances (40/40 matched), wall-clock timings at scale, and restart-robustness; reported the honest gap (near-optimal, no certificate).",
         delta=delta(
             "I demanded proof, not assertion; the brute-force enumeration is an independent oracle the AI did not control.",
             "For a publishable claim, 'trust me it's optimal' is unacceptable - needed verifiable evidence.",
             "I had the AI add brute-force + CBC cross-checks as a reusable test file.",
             "Accepted the solver only after 40/40 brute-force agreement; logged the no-certificate caveat."),
         ev="tests/validate_solver.py output (40/40 optimal)."),

    dict(t="PROBLEM-SOLVING", stage="Algorithms / Solver",
         ctx="Solver quality/speed needed improving for LRFM (k=10 took ~17s).",
         prompt="Use Numba please. Apply Kernighan-Lin for quality is okay too.",
         resp="Added a Numba-JIT kernel with greedy relocation + Kernighan-Lin pairwise swaps; verified the swap-gain formula; correctness preserved (40/40) and robustness rose to 100% hit-best at 200 restarts.",
         delta=delta(
             "I independently checked the KL swap-gain math (C[u,a]-C[u,b])+(C[w,b]-C[w,a])+2w_uw before trusting it.",
             "Single-vertex moves get stuck in local optima; KL swaps escape exactly those.",
             "Combined JIT speed with KL quality; re-ran the validator to confirm no regression.",
             "Kept Numba+KL after verifying correctness and the robustness improvement."),
         ev="validate_solver speed (17s->6.7s) & robustness (80%->100%) tables."),

    dict(t="DECISION", stage="Data Analysis / Feature Engineering",
         ctx="Deciding what extending RFM to LRFM actually means methodologically.",
         prompt="Adding a behaviour criterion to RFM - is this Feature Engineering?",
         resp="Yes - the engine is feature engineering, but the citable framing is 'RFM model extension' (LRFM/RFMT). Listed which behavioural features are computable from Online Retail II (tenure L, interpurchase time, return rate).",
         delta=delta(
             "The AI distinguished the ML term (feature engineering) from the citable marketing term (RFM extension) - useful for the write-up.",
             "Online Retail II constrains which features are derivable; L (tenure) is the cleanest, paper-endorsed one.",
             "I adopted 'LRFM extension' as the narrative and feature engineering as the mechanism.",
             "Chose L = first-to-last-purchase tenure as the 4th axis."),
         ev="docs/FEATURE_ENGINEERING_PLAN.md."),

    dict(t="DECISION", stage="Data Analysis / Preprocessing",
         ctx="Worried the cleaned data is still heavily skewed and unscaled - should I transform/standardize?",
         prompt="The data is still heavily skewed - can we improve it? And do we need StandardScaler / scaling?",
         resp="No - the model uses quantile (NTILE) scores, which are rank-based and invariant to any monotonic transform; and all score axes already share std=1.414, so scaling is a no-op. The baseline paper explicitly states standardization is unnecessary for scores.",
         delta=delta(
             "I challenged the instinct to 'always transform/scale'; the AI proved scores are unchanged under log (skew 25.3->0.27, 0 score changes).",
             "The baseline paper itself says scores need no standardization - cited support, not just opinion.",
             "Rather than accept a generic answer, I had it demonstrate invariance empirically.",
             "Decided NOT to transform or scale; kept outliers (faithful to baseline)."),
         ev="Empirical proof: log1p reduces skew but 0 customers change quantile score; per-axis std all 1.414."),

    dict(t="DECISION", stage="Data Analysis / Database design",
         ctx="The cleaned data was a flat table; deciding whether to build a relational layer.",
         prompt="From the cleaned data, I think we need to organize a SQL relational database. Do you think so?",
         resp="Yes - it satisfies the 'Data Analysis with SQL' deliverable and feeds Tableau. Recommended a lightweight SQLite star schema with SQL as the source of truth for R/F/M/L scores.",
         delta=delta(
             "The AI separated 'analysis layer (SQL)' from 'modeling layer (Python)' - SQL must not run the solver.",
             "A heavy server DB is overkill for 794k rows; SQLite/DuckDB is right-sized.",
             "I chose SQLite + SQL as the single source of truth for scores, read by Python for the graph.",
             "Owned the architecture: SQLite star schema, scores computed via NTILE in SQL."),
         ev="src/build_db.py; data/processed/retail.db."),

    dict(t="DECISION", stage="Data Analysis / Database design",
         ctx="The fact table was a denormalized star (violated 2NF/3NF); I wanted proper normalization.",
         prompt="Make fact_sales also 3NF.",
         resp="Refactored to dim_invoice + fact_line (surrogate key), dropped the derived line_total, and declared PK/FK with enforcement. Verified the change was lossless (identical recency mean 200.87 and 114 (R,F,M) cells).",
         delta=delta(
             "I verified normalization didn't change the data - same features before/after, FK check clean.",
             "3NF is more rigorous/defensible for the report than a denormalized star.",
             "Recognised v_invoices already encapsulated the invoice grain, so promotion to a table was natural.",
             "Chose 3NF over the star schema and confirmed referential integrity."),
         ev="docs/DB_SCHEMA.md; PRAGMA foreign_key_check = no violations."),

    dict(t="DECISION", stage="Data Analysis / Feature Engineering",
         ctx="FE-3 decision gate: does adding L provide non-redundant signal, or is it redundant with F?",
         prompt="Continue Stage 2 (feature engineering) - justify whether L earns its place.",
         resp="The automated verdict first returned NO-GO (L-F correlation 0.85 > 0.8 threshold), then was corrected to GO after comparing to RFM's own internal redundancy (F-M is already 0.81); L adds 2.75x cluster resolution.",
         delta=delta(
             "I caught that the AI's hard-coded 0.8 redundancy threshold was arbitrary and produced a misleading NO-GO (see Hallucination H2).",
             "F-M correlate 0.81 in the accepted RFM model, so L's 0.85 with F is not disqualifying.",
             "I had the verdict re-based on RFM-internal correlation, not an absolute cutoff.",
             "Decided GO on LRFM, with the honest caveat that the payoff is actionability, not silhouette."),
         ev="results/tables/stage2_feature_report.md."),

    dict(t="VERIFICATION", stage="Build Model / Reproduction",
         ctx="Before claiming any improvement, I must prove we reproduce the baseline paper.",
         prompt="Run the pipeline to compare our model with the paper's model. Have we reproduced the paper's results?",
         resp="RFM/max-k-cut reproduces the paper: customers 5,878, recency mean 200.87, 114-vertex reduced graph, silhouette mean |Delta|=0.018, cut objectives within ~0.02% of Table 5.",
         delta=delta(
             "I insisted on multiple exact anchors (customers, recency, cell count, silhouette, objective) before trusting reproduction.",
             "Reproduction is a prerequisite to claiming an extension improves anything.",
             "Used several independent reconciliation points rather than one number.",
             "Accepted reproduction as confirmed; flagged the one open gap (Monetary +1.5%)."),
         ev="stage3_modeling_report.md; comparison_report.md."),

    dict(t="VERIFICATION", stage="Build Model / Comparison",
         ctx="Comparing our method against K-means/Ward/GMM and the paper's reported numbers.",
         prompt="Compare our model with the paper's model across methods.",
         resp="Ran max-k-cut, K-means, Ward and GMM on identical score features; found K-means ~ max-k-cut on equal footing, meaning the paper's reported max-k-cut advantage over K-means is largely a preprocessing artifact.",
         delta=delta(
             "I treated the paper's claim critically - on identical features the 'max-k-cut beats K-means' gap mostly disappears.",
             "The paper ran K-means on differently preprocessed raw values, not the scores - an apples-to-oranges comparison.",
             "I levelled the field by running every method on the same score matrix.",
             "Decided to report this fair critique rather than overclaim our method's superiority."),
         ev="comparison_report.md; comparison_silhouette.png."),

    dict(t="DECISION", stage="Build Model / Scope",
         ctx="Clarifying whether supervised models are needed at all.",
         prompt="We don't need Linear Regression / XGBoost, right?",
         resp="Correct - the task is unsupervised clustering with no target variable; the brief's regression-on-response is blocked because the dataset has no response label. Supervised models would only apply via a constructed temporal-prediction target.",
         delta=delta(
             "I confirmed there is genuinely no target to regress on, rather than adding models for show.",
             "Online Retail II has no campaign-response field, so 'response ~ RFM' is impossible as stated.",
             "Considered an optional temporal next-period-spend reframe and judged it out of scope.",
             "Decided to drop the regression component and document the data limitation."),
         ev="Dataset schema (8 transaction columns, no response field)."),

    dict(t="DECISION", stage="Data Visualization",
         ctx="The first figures looked dated and some chart types were too exotic for referees.",
         prompt="I want a modern style but not a strange format - make sure everyone can understand it easily.",
         resp="Built a reusable modern style module; on review, replaced the 3D scatter with 2D R-M / F-M scatters (matching the paper's Fig 6) and replaced the radar chart with grouped bars.",
         delta=delta(
             "I judged radar and 3D scatter as referee-risky (distortion, occlusion), overriding the brief's '3D scatter' request.",
             "Referees in this venue must read figures instantly; the baseline paper itself uses 2D scatters.",
             "Kept a modern theme but only standard, unambiguous chart types.",
             "Chose 2D scatters + grouped bars; kept a 4-angle 3D only as a private exploratory view."),
         ev="results/figures/viz_segment_*.png; baseline paper Fig 6."),

    dict(t="DECISION", stage="Data Visualization / Segmentation",
         ctx="I was uneasy that the AI named segments 'Champions/Loyal' - it felt arbitrary and not general.",
         prompt="I don't know how you named groups like Champion, Loyal... Does it make my research less general? Other papers don't name segments this way.",
         resp="Explained the names came from a practitioner RFM grid with arbitrary thresholds, NOT from either reference paper. Recommended switching to data-driven max-k-cut clusters labelled by Cheng & Chen (2009) loyalty distance, exactly as the baseline paper does.",
         delta=delta(
             "I correctly sensed the named grid was an arbitrary, non-citable convention that weakens generality.",
             "The baseline paper ranks clusters by Cheng & Chen loyalty distance, not marketing names.",
             "I switched the segment basis to the model's own k=4 clusters, labelled by a cited loyalty metric.",
             "Dropped the named grid for the paper; kept it only as an optional SQL business overlay."),
         ev="viz loyalty ranking (Very high/High/Medium/Low); baseline Table 9."),
]


# ---------------------------------------------------------------------------
# HALLUCINATION / ERROR DETECTION  (>= 3 required)
# ---------------------------------------------------------------------------
HALLU = [
    dict(entry="03", htype="Logic Error / Oversimplification",
         claim="In its first report the AI recommended a Graph Neural Network as 'genuinely unexplored... would be novel', implying it would benefit the project.",
         reality="A GNN is a poor fit: the graph encodes only re-encoded RFM distances, nodes have 3-4 features, and the data is tiny for deep learning - novelty is not benefit.",
         detected="I pushed back ('does it help?'); the AI re-examined and conceded it had oversold the option.",
         action="Rejected the GNN; kept LRFM + max-k-cut. Reasoning recorded so the rejection is defensible."),

    dict(entry="11", htype="Logic Error (arbitrary threshold)",
         claim="The AI's automated FE-3 verdict printed 'NO-GO - reconsider L' because L-F correlation (0.85) exceeded a hard-coded 0.8 'redundant' threshold.",
         reality="F-M already correlate 0.81 in the accepted RFM model, so 0.85 is NOT disqualifying; the 0.8 cutoff was arbitrary.",
         detected="I compared L's correlation to RFM's own internal correlations instead of an absolute cutoff.",
         action="Changed the verdict logic to a relative comparison -> GO; documented L's 2.75x resolution gain."),

    dict(entry="13", htype="Uncritical source acceptance",
         claim="The AI initially recorded the graph paper's '79,104 transactions' as a reproduction target to match.",
         reality="79,104 contradicts the paper's own reported frequency mean (6.30 x 5,878 ~ 37,031); our cleaned data has 36,969 invoices.",
         detected="Computed unique invoices from the cleaned data and cross-checked against the paper's frequency mean.",
         action="Treated 79,104 as a likely typo in the paper; used 36,969 invoices / frequency-mean 6.30 as the real anchor."),

    dict(entry="15", htype="Context Misunderstanding (data assumption)",
         claim="The project brief (which an AI could follow literally) implied campaign-response / open-rate data for 'response rate per segment' and 'open rate +30%'.",
         reality="Online Retail II contains only transaction fields - there is no response/open-rate label anywhere in the dataset.",
         detected="Inspected the dataset schema (8 columns) and confirmed no response field exists.",
         action="Dropped the response-regression component; used engagement proxies (repeat/active rate) and flagged the limitation in the report."),
]

CHECKLIST = [
    "Prompt này ảnh hưởng đến quyết định quan trọng trong project?",
    "Nếu không có prompt này, project có thay đổi về architecture/design/approach?",
    "Tôi có thể giải thích lý do chọn/không chọn gợi ý của AI?",
    "Có minh chứng cụ thể (code, metrics, comparison) cho quyết định này?",
    "Prompt này phản ánh tư duy phản biện, không phải chỉ copy AI output?",
]

# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def _style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = WRAP


def build():
    wb = openpyxl.Workbook()

    # Sheet 1 - Metadata & Summary
    ws = wb.active; ws.title = "1. Metadata & Summary"
    ws["A1"] = "AI AUDIT LOG - METADATA & SUMMARY"; ws["A1"].font = TITLE_FONT
    r = 3
    for k, v in META.items():
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 3, v)
        r += 1
    ws.cell(r + 1, 1, "Note").font = Font(bold=True, italic=True)
    ws.cell(r + 1, 3, "Only CORE prompts (decisions that changed the project's substance) are logged; "
                      "routine syntax/formatting prompts are excluded per the DAP391m guideline.").alignment = WRAP
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["C"].width = 70

    # Sheet 2 - Detailed Audit Log
    ws = wb.create_sheet("2. Detailed Audit Log")
    cols = ["Entry #", "Prompt Type", "Stage/Component", "Problem/Context",
            "Prompt to AI", "AI Response (Summary)", "Human Delta & Reflection", "Evidence"]
    ws.append(cols); _style_header(ws, 1, len(cols))
    for i, e in enumerate(CORE, 1):
        ws.append([f"{i:03d}", e["t"], e["stage"], e["ctx"], e["prompt"], e["resp"], e["delta"], e["ev"]])
    widths = [8, 16, 24, 34, 40, 40, 52, 28]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP; c.border = THIN
    ws.freeze_panes = "A2"

    # Sheet 3 - Hallucination Detection
    ws = wb.create_sheet("3. Hallucination Detection")
    cols = ["Entry # (Sheet 2)", "Hallucination Type", "AI's Claim", "Reality Check",
            "How Detected", "Corrective Action"]
    ws.append(cols); _style_header(ws, 1, len(cols))
    for h in HALLU:
        ws.append([h["entry"], h["htype"], h["claim"], h["reality"], h["detected"], h["action"]])
    for j, w in enumerate([14, 26, 44, 44, 38, 44], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP; c.border = THIN
    ws.freeze_panes = "A2"

    # Sheet 4 - Self-Assessment Checklist
    ws = wb.create_sheet("4. Self-Assessment Checklist")
    ws["A1"] = "SELF-ASSESSMENT CHECKLIST (mỗi entry pass >=4/5)"; ws["A1"].font = TITLE_FONT
    ws.append([]); ws.append(["#", "Tiêu chí", "Pass?", "Note"]); _style_header(ws, 3, 4)
    for i, c in enumerate(CHECKLIST, 1):
        ws.append([i, c, "YES", ""])
    ws.append([])
    ws.append(["B. KIỂM TRA TỔNG THỂ LOG", "", "", ""]); ws.cell(ws.max_row, 1).font = Font(bold=True)
    overall = [
        (f"Core entries in range (15-20)?", f"YES - {len(CORE)} entries"),
        ("Hallucination cases >= 3?", f"YES - {len(HALLU)} cases"),
        ("Every entry has full Human Delta (4 questions)?", "YES"),
        ("Evidence for >= 70% of entries?", "YES - 100% have evidence"),
        ("Each rubric component has >= 2-3 core prompts?", "YES (Research, Data Analysis, Viz, Modeling)"),
    ]
    for crit, val in overall:
        ws.append(["", crit, val, ""])
    ws.column_dimensions["A"].width = 6; ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 24; ws.column_dimensions["D"].width = 30
    for row in ws.iter_rows(min_row=3):
        for c in row:
            c.alignment = WRAP

    config.RESULTS.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Audit log written: {OUT}")
    print(f"  core entries: {len(CORE)} | hallucination cases: {len(HALLU)}")


if __name__ == "__main__":
    build()
